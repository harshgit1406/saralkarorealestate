import json
import re
from io import BytesIO
from typing import Any

import asyncpg
from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from starlette.responses import Response, StreamingResponse

from app.api.dependencies import get_current_user, get_db_pool
from app.api.v1.crud import audit_log, clean_payload, ensure_permission, json_ready

router = APIRouter()

ENTITY_FIELDS = [
    "project_id",
    "parent_id",
    "entity_type",
    "entity_code",
    "name",
    "inventory_status",
    "lifecycle_stage",
    "level_no",
    "path",
    "sort_order",
    "metadata",
]
DIMENSION_FIELDS = [
    "area",
    "carpet_area",
    "builtup_area",
    "saleable_area",
    "length",
    "width",
    "measurement_unit",
]
PRICING_FIELDS = ["base_price", "final_price", "price_per_sqft", "currency", "pricing_metadata"]
DETAIL_FIELDS = ["facing", "bhk_type", "display_note", "notes", "metadata"]
MAP_FIELDS = ["project_id", "map_name", "map_engine", "version_no", "is_published", "map_data", "thumbnail_url"]
ELEMENT_FIELDS = ["project_map_id", "inventory_entity_id", "element_id", "element_type", "is_interactive", "metadata"]
PROJECT_WITH_MAP_FIELDS = ["name", "project_code", "project_type", "location", "description", "status"]
EXPORT_HEADERS = [
    "Entity Code",
    "Parent Code",
    "Type",
    "Name",
    "Status",
    "Area",
    "Price",
    "Facing",
    "Display Note",
]
VALID_IMPORT_TYPES = {"floor", "flat", "shop", "office", "villa", "parking", "other"}
VALID_IMPORT_STATUSES = {"available", "blocked", "booked", "sold", "hold", "reserved", "inactive"}


def extract_svg_plot_ids(svg: str) -> list[str]:
    ids = re.findall(r"""<rect\b[^>]*\bid=["']([^"']+)["']""", svg, flags=re.IGNORECASE)
    seen: set[str] = set()
    plot_ids = []
    for element_id in ids:
        if element_id.lower().startswith("plot_") and element_id not in seen:
            seen.add(element_id)
            plot_ids.append(element_id)
    return plot_ids


def extract_view_box(svg: str) -> str | None:
    match = re.search(r"""\bviewBox=["']([^"']+)["']""", svg, flags=re.IGNORECASE)
    return match.group(1) if match else None


def clean_cell(value: Any) -> str:
    return str(value or "").strip()


async def assert_project(
    connection: asyncpg.Connection,
    organization_id: int,
    project_id: int,
) -> None:
    exists = await connection.fetchval(
        "SELECT EXISTS (SELECT 1 FROM projects WHERE id = $1 AND organization_id = $2)",
        project_id,
        organization_id,
    )
    if not exists:
        raise HTTPException(status_code=422, detail="Invalid project_id")


async def upsert_one_to_one(
    connection: asyncpg.Connection,
    *,
    table: str,
    organization_id: int,
    inventory_entity_id: int,
    payload: dict[str, Any],
    fields: list[str],
) -> dict[str, Any] | None:
    data = clean_payload(payload, fields)
    if not data:
        return None
    columns = ["organization_id", "inventory_entity_id", *data.keys()]
    placeholders = [f"${index}" for index in range(1, len(columns) + 1)]
    update_columns = [f"{column} = EXCLUDED.{column}" for column in data.keys()]
    row = await connection.fetchrow(
        f"""
        INSERT INTO {table} ({", ".join(columns)})
        VALUES ({", ".join(placeholders)})
        ON CONFLICT (inventory_entity_id) DO UPDATE
        SET {", ".join(update_columns)}
        RETURNING *
        """,
        organization_id,
        inventory_entity_id,
        *data.values(),
    )
    return json_ready(row)


async def fetch_inventory_detail(
    connection: asyncpg.Connection,
    organization_id: int,
    entity_id: int,
) -> dict[str, Any]:
    row = await connection.fetchrow(
        """
        SELECT
            ie.*,
            p.name AS project_name,
            row_to_json(dim) AS dimensions,
            row_to_json(pr) AS pricing,
            row_to_json(det) AS details,
            (
                SELECT row_to_json(bdata)
                FROM (
                    SELECT
                        b.*,
                        c.full_name AS customer_name,
                        c.phone AS customer_phone,
                        l.name AS lead_name
                    FROM bookings b
                    LEFT JOIN booking_applicants ba ON ba.booking_id = b.id AND ba.is_primary = TRUE
                    LEFT JOIN customers c ON c.id = ba.customer_id
                    LEFT JOIN leads l ON l.id = b.lead_id
                    WHERE b.inventory_entity_id = ie.id
                      AND b.booking_status IN ('reserved', 'confirmed')
                    ORDER BY b.created_at DESC
                    LIMIT 1
                ) bdata
            ) AS active_booking
        FROM inventory_entities ie
        JOIN projects p ON p.id = ie.project_id
        LEFT JOIN inventory_dimensions dim ON dim.inventory_entity_id = ie.id
        LEFT JOIN inventory_pricing pr ON pr.inventory_entity_id = ie.id
        LEFT JOIN inventory_details det ON det.inventory_entity_id = ie.id
        WHERE ie.id = $1 AND ie.organization_id = $2
        """,
        entity_id,
        organization_id,
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Inventory entity not found")
    return json_ready(row)


@router.get("/entities")
async def list_entities(
    project_id: int | None = None,
    parent_id: int | None = None,
    entity_type: str | None = None,
    status_filter: str | None = Query(default=None, alias="status"),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.view")
        rows = await connection.fetch(
            """
            SELECT
                ie.*,
                p.name AS project_name,
                dim.saleable_area,
                dim.carpet_area,
                pr.final_price,
                pr.price_per_sqft,
                det.facing,
                det.bhk_type,
                det.display_note
            FROM inventory_entities ie
            JOIN projects p ON p.id = ie.project_id
            LEFT JOIN inventory_dimensions dim ON dim.inventory_entity_id = ie.id
            LEFT JOIN inventory_pricing pr ON pr.inventory_entity_id = ie.id
            LEFT JOIN inventory_details det ON det.inventory_entity_id = ie.id
            WHERE ie.organization_id = $1
              AND ($2::int IS NULL OR ie.project_id = $2)
              AND ($3::int IS NULL OR ie.parent_id = $3)
              AND ($4::varchar IS NULL OR ie.entity_type = $4)
              AND ($5::varchar IS NULL OR ie.inventory_status = $5)
            ORDER BY ie.path, ie.sort_order, ie.entity_code
            """,
            current_user["organization_id"],
            project_id,
            parent_id,
            entity_type,
            status_filter,
        )
    return {"items": json_ready(rows)}


@router.get("/entities/{entity_id}")
async def get_entity(
    entity_id: int,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.view")
        return await fetch_inventory_detail(connection, current_user["organization_id"], entity_id)


@router.post("/entities", status_code=status.HTTP_201_CREATED)
async def create_entity(
    payload: dict[str, Any] = Body(...),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.create")
        await assert_project(connection, current_user["organization_id"], int(payload["project_id"]))
        async with connection.transaction():
            entity_data = clean_payload(payload, ENTITY_FIELDS)
            project = await connection.fetchrow(
                "SELECT project_code FROM projects WHERE id = $1 AND organization_id = $2",
                int(payload["project_id"]),
                current_user["organization_id"],
            )
            if not entity_data.get("entity_code"):
                entity_type = str(entity_data.get("entity_type") or "unit").upper()
                prefix = {
                    "PLOT": "PLT",
                    "FLOOR": "FLR",
                    "FLAT": "FLT",
                    "SHOP": "SHP",
                    "OFFICE": "OFC",
                    "VILLA": "VIL",
                }.get(entity_type, "UNT")
                count = await connection.fetchval(
                    """
                    SELECT COUNT(*) + 1
                    FROM inventory_entities
                    WHERE organization_id = $1 AND project_id = $2 AND entity_type = $3
                    """,
                    current_user["organization_id"],
                    int(payload["project_id"]),
                    entity_data.get("entity_type"),
                )
                entity_data["entity_code"] = f"{prefix}-{int(count):03d}"
            if not entity_data.get("path"):
                parent_path = None
                if entity_data.get("parent_id"):
                    parent_path = await connection.fetchval(
                        """
                        SELECT path FROM inventory_entities
                        WHERE id = $1 AND organization_id = $2 AND project_id = $3
                        """,
                        int(entity_data["parent_id"]),
                        current_user["organization_id"],
                        int(payload["project_id"]),
                    )
                root_path = parent_path or (project["project_code"] if project and project["project_code"] else f"PRJ-{payload['project_id']}")
                entity_data["path"] = f"{root_path}/{entity_data['entity_code']}"
            columns = ["organization_id", *entity_data.keys()]
            row = await connection.fetchrow(
                f"""
                INSERT INTO inventory_entities ({", ".join(columns)})
                VALUES ({", ".join(f"${i}" for i in range(1, len(columns) + 1))})
                RETURNING *
                """,
                current_user["organization_id"],
                *entity_data.values(),
            )
            await upsert_one_to_one(
                connection,
                table="inventory_dimensions",
                organization_id=current_user["organization_id"],
                inventory_entity_id=row["id"],
                payload=payload.get("dimensions", {}),
                fields=DIMENSION_FIELDS,
            )
            await upsert_one_to_one(
                connection,
                table="inventory_pricing",
                organization_id=current_user["organization_id"],
                inventory_entity_id=row["id"],
                payload=payload.get("pricing", {}),
                fields=PRICING_FIELDS,
            )
            await upsert_one_to_one(
                connection,
                table="inventory_details",
                organization_id=current_user["organization_id"],
                inventory_entity_id=row["id"],
                payload=payload.get("details", {}),
                fields=DETAIL_FIELDS,
            )
            detail = await fetch_inventory_detail(connection, current_user["organization_id"], row["id"])
            await audit_log(
                connection,
                organization_id=current_user["organization_id"],
                user_id=current_user["id"],
                entity_type="inventory",
                entity_id=row["id"],
                action="created",
                new_data=detail,
            )
    return detail


@router.patch("/entities/{entity_id}")
async def update_entity(
    entity_id: int,
    payload: dict[str, Any] = Body(...),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.update")
        old = await fetch_inventory_detail(connection, current_user["organization_id"], entity_id)
        async with connection.transaction():
            entity_data = clean_payload(payload, ENTITY_FIELDS)
            if entity_data:
                assignments = [
                    f"{column} = ${index}" for index, column in enumerate(entity_data.keys(), start=1)
                ]
                await connection.execute(
                    f"""
                    UPDATE inventory_entities
                    SET {", ".join(assignments)}
                    WHERE id = ${len(entity_data) + 1} AND organization_id = ${len(entity_data) + 2}
                    """,
                    *entity_data.values(),
                    entity_id,
                    current_user["organization_id"],
                )
            await upsert_one_to_one(
                connection,
                table="inventory_dimensions",
                organization_id=current_user["organization_id"],
                inventory_entity_id=entity_id,
                payload=payload.get("dimensions", {}),
                fields=DIMENSION_FIELDS,
            )
            await upsert_one_to_one(
                connection,
                table="inventory_pricing",
                organization_id=current_user["organization_id"],
                inventory_entity_id=entity_id,
                payload=payload.get("pricing", {}),
                fields=PRICING_FIELDS,
            )
            await upsert_one_to_one(
                connection,
                table="inventory_details",
                organization_id=current_user["organization_id"],
                inventory_entity_id=entity_id,
                payload=payload.get("details", {}),
                fields=DETAIL_FIELDS,
            )
            detail = await fetch_inventory_detail(connection, current_user["organization_id"], entity_id)
            await audit_log(
                connection,
                organization_id=current_user["organization_id"],
                user_id=current_user["id"],
                entity_type="inventory",
                entity_id=entity_id,
                action="updated",
                old_data=old,
                new_data=detail,
            )
    return detail


@router.delete("/entities/{entity_id}")
async def delete_entity(
    entity_id: int,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.delete")
        old = await fetch_inventory_detail(connection, current_user["organization_id"], entity_id)
        deleted = await connection.fetchrow(
            "DELETE FROM inventory_entities WHERE id = $1 AND organization_id = $2 RETURNING id",
            entity_id,
            current_user["organization_id"],
        )
        if deleted is None:
            raise HTTPException(status_code=404, detail="Inventory entity not found")
        await audit_log(
            connection,
            organization_id=current_user["organization_id"],
            user_id=current_user["id"],
            entity_type="inventory",
            entity_id=entity_id,
            action="deleted",
            old_data=old,
        )
    return {"deleted": True, "id": entity_id}


@router.get("/tree")
async def inventory_tree(
    project_id: int,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.view")
        await assert_project(connection, current_user["organization_id"], project_id)
        rows = await connection.fetch(
            """
            SELECT id, parent_id, entity_type, entity_code, name, inventory_status, level_no, path, sort_order
            FROM inventory_entities
            WHERE organization_id = $1 AND project_id = $2
            ORDER BY level_no, sort_order, entity_code
            """,
            current_user["organization_id"],
            project_id,
        )
    return {"items": json_ready(rows)}


@router.get("/excel")
async def export_inventory_excel(
    project_id: int,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> Response:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.view")
        project = await connection.fetchrow(
            "SELECT name, project_code FROM projects WHERE id = $1 AND organization_id = $2",
            project_id,
            current_user["organization_id"],
        )
        if project is None:
            raise HTTPException(status_code=404, detail="Project not found")
        rows = await connection.fetch(
            """
            SELECT
                ie.entity_code,
                parent.entity_code AS parent_code,
                ie.entity_type,
                ie.name,
                ie.inventory_status,
                dim.saleable_area,
                pr.final_price,
                det.facing,
                det.display_note
            FROM inventory_entities ie
            LEFT JOIN inventory_entities parent ON parent.id = ie.parent_id
            LEFT JOIN inventory_dimensions dim ON dim.inventory_entity_id = ie.id
            LEFT JOIN inventory_pricing pr ON pr.inventory_entity_id = ie.id
            LEFT JOIN inventory_details det ON det.inventory_entity_id = ie.id
            WHERE ie.organization_id = $1 AND ie.project_id = $2
            ORDER BY ie.path, ie.sort_order, ie.entity_code
            """,
            current_user["organization_id"],
            project_id,
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="13265C")

    for row in rows:
        sheet.append(
            [
                row["entity_code"],
                row["parent_code"] or "",
                row["entity_type"],
                row["name"],
                row["inventory_status"],
                float(row["saleable_area"] or 0),
                float(row["final_price"] or 0),
                row["facing"] or "",
                row["display_note"] or "",
            ]
        )

    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(sorted(VALID_IMPORT_STATUSES))}"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(sorted(VALID_IMPORT_TYPES | {"plot"}))}"',
        allow_blank=False,
    )
    sheet.add_data_validation(status_validation)
    sheet.add_data_validation(type_validation)
    status_validation.add("E2:E1000")
    type_validation.add("C2:C1000")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(sheet.max_row, 2)}"
    widths = [18, 18, 14, 28, 16, 12, 14, 16, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "How to edit"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A3"] = "Do not add or edit database IDs. This workbook does not contain IDs."
    instructions["A4"] = "Edit existing rows by changing Name, Status, Area, Price, Facing, or Display Note."
    instructions["A5"] = "To add a new child row, leave Entity Code blank, set Parent Code to an existing plot/floor code, choose Type, and fill Name."
    instructions["A6"] = "Plots should be created from the SVG map upload/update flow, not from this sheet."
    instructions.column_dimensions["A"].width = 120

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", project["project_code"] or project["name"]).strip("_")
    filename = f"{safe_name or 'inventory'}_inventory.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/excel/import")
async def import_inventory_excel(
    project_id: int,
    request: Request,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    content = await request.body()
    if not content:
        raise HTTPException(status_code=422, detail="Excel file body is required")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Could not read Excel file") from exc
    if "Inventory" not in workbook.sheetnames:
        raise HTTPException(status_code=422, detail="Inventory sheet is required")

    sheet = workbook["Inventory"]
    headers = [clean_cell(cell.value) for cell in sheet[1]]
    missing = [header for header in EXPORT_HEADERS if header not in headers]
    if missing:
        raise HTTPException(status_code=422, detail=f"Missing columns: {', '.join(missing)}")
    indexes = {header: headers.index(header) + 1 for header in EXPORT_HEADERS}

    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.update")
        await ensure_permission(connection, current_user, "inventory.create")
        await assert_project(connection, current_user["organization_id"], project_id)
        existing = await connection.fetch(
            """
            SELECT id, entity_code, entity_type, level_no, path
            FROM inventory_entities
            WHERE organization_id = $1 AND project_id = $2
            """,
            current_user["organization_id"],
            project_id,
        )
        by_code = {row["entity_code"]: row for row in existing}
        updated = 0
        created = 0
        skipped = 0
        errors: list[str] = []

        async with connection.transaction():
            for row_no in range(2, sheet.max_row + 1):
                entity_code = clean_cell(sheet.cell(row_no, indexes["Entity Code"]).value)
                parent_code = clean_cell(sheet.cell(row_no, indexes["Parent Code"]).value)
                entity_type = clean_cell(sheet.cell(row_no, indexes["Type"]).value).lower()
                name = clean_cell(sheet.cell(row_no, indexes["Name"]).value)
                status_value = clean_cell(sheet.cell(row_no, indexes["Status"]).value).lower() or "available"
                area = sheet.cell(row_no, indexes["Area"]).value or 0
                price = sheet.cell(row_no, indexes["Price"]).value or 0
                facing = clean_cell(sheet.cell(row_no, indexes["Facing"]).value)
                display_note = clean_cell(sheet.cell(row_no, indexes["Display Note"]).value)

                if not any([entity_code, parent_code, entity_type, name]):
                    skipped += 1
                    continue
                if status_value not in VALID_IMPORT_STATUSES:
                    errors.append(f"Row {row_no}: invalid status")
                    continue
                if entity_code:
                    current = by_code.get(entity_code)
                    if current is None:
                        errors.append(f"Row {row_no}: entity code not found")
                        continue
                    if not name:
                        errors.append(f"Row {row_no}: name is required")
                        continue
                    await connection.execute(
                        """
                        UPDATE inventory_entities
                        SET name = $1, inventory_status = $2
                        WHERE id = $3 AND organization_id = $4
                        """,
                        name,
                        status_value,
                        current["id"],
                        current_user["organization_id"],
                    )
                    await upsert_one_to_one(
                        connection,
                        table="inventory_dimensions",
                        organization_id=current_user["organization_id"],
                        inventory_entity_id=current["id"],
                        payload={"saleable_area": float(area or 0)},
                        fields=DIMENSION_FIELDS,
                    )
                    await upsert_one_to_one(
                        connection,
                        table="inventory_pricing",
                        organization_id=current_user["organization_id"],
                        inventory_entity_id=current["id"],
                        payload={"final_price": float(price or 0), "currency": "INR"},
                        fields=PRICING_FIELDS,
                    )
                    await upsert_one_to_one(
                        connection,
                        table="inventory_details",
                        organization_id=current_user["organization_id"],
                        inventory_entity_id=current["id"],
                        payload={"facing": facing or None, "display_note": display_note or None},
                        fields=DETAIL_FIELDS,
                    )
                    updated += 1
                    continue

                if entity_type == "plot":
                    errors.append(f"Row {row_no}: plots must be created from SVG map upload")
                    continue
                if entity_type not in VALID_IMPORT_TYPES:
                    errors.append(f"Row {row_no}: invalid type")
                    continue
                parent = by_code.get(parent_code)
                if parent is None:
                    errors.append(f"Row {row_no}: parent code not found")
                    continue
                if not name:
                    errors.append(f"Row {row_no}: name is required")
                    continue
                payload = {
                    "project_id": project_id,
                    "parent_id": parent["id"],
                    "entity_type": entity_type,
                    "name": name,
                    "inventory_status": status_value,
                    "lifecycle_stage": "active_sales",
                    "level_no": int(parent["level_no"]) + 1,
                    "sort_order": 99,
                    "dimensions": {"saleable_area": float(area or 0), "measurement_unit": "sqft"},
                    "pricing": {"final_price": float(price or 0), "currency": "INR"},
                    "details": {"facing": facing or None, "display_note": display_note or None},
                }
                entity_data = clean_payload(payload, ENTITY_FIELDS)
                count = await connection.fetchval(
                    """
                    SELECT COUNT(*) + 1
                    FROM inventory_entities
                    WHERE organization_id = $1 AND project_id = $2 AND entity_type = $3
                    """,
                    current_user["organization_id"],
                    project_id,
                    entity_type,
                )
                prefix = {"floor": "FLR", "flat": "FLT", "shop": "SHP", "office": "OFC", "villa": "VIL"}.get(entity_type, "UNT")
                entity_data["entity_code"] = f"{prefix}-{int(count):03d}"
                entity_data["path"] = f"{parent['path']}/{entity_data['entity_code']}"
                columns = ["organization_id", *entity_data.keys()]
                entity = await connection.fetchrow(
                    f"""
                    INSERT INTO inventory_entities ({", ".join(columns)})
                    VALUES ({", ".join(f"${i}" for i in range(1, len(columns) + 1))})
                    RETURNING *
                    """,
                    current_user["organization_id"],
                    *entity_data.values(),
                )
                by_code[entity["entity_code"]] = entity
                await upsert_one_to_one(
                    connection,
                    table="inventory_dimensions",
                    organization_id=current_user["organization_id"],
                    inventory_entity_id=entity["id"],
                    payload=payload["dimensions"],
                    fields=DIMENSION_FIELDS,
                )
                await upsert_one_to_one(
                    connection,
                    table="inventory_pricing",
                    organization_id=current_user["organization_id"],
                    inventory_entity_id=entity["id"],
                    payload=payload["pricing"],
                    fields=PRICING_FIELDS,
                )
                await upsert_one_to_one(
                    connection,
                    table="inventory_details",
                    organization_id=current_user["organization_id"],
                    inventory_entity_id=entity["id"],
                    payload=payload["details"],
                    fields=DETAIL_FIELDS,
                )
                created += 1

    if errors:
        return {"updated": updated, "created": created, "skipped": skipped, "errors": errors[:25]}
    return {"updated": updated, "created": created, "skipped": skipped, "errors": []}


@router.get("/map")
async def inventory_map(
    project_id: int,
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "inventory.view")
        project = await connection.fetchrow(
            "SELECT * FROM projects WHERE id = $1 AND organization_id = $2",
            project_id,
            current_user["organization_id"],
        )
        if project is None:
            raise HTTPException(status_code=404, detail="Project not found")
        project_map = await connection.fetchrow(
            """
            SELECT * FROM project_maps
            WHERE project_id = $1 AND organization_id = $2 AND is_published = TRUE
            ORDER BY version_no DESC, id DESC
            LIMIT 1
            """,
            project_id,
            current_user["organization_id"],
        )
        elements = await connection.fetch(
            """
            SELECT
                me.*,
                ie.entity_code,
                ie.name AS inventory_name,
                ie.entity_type AS inventory_type,
                ie.inventory_status,
                pr.final_price,
                dim.saleable_area
            FROM map_elements me
            LEFT JOIN inventory_entities ie ON ie.id = me.inventory_entity_id
            LEFT JOIN inventory_pricing pr ON pr.inventory_entity_id = ie.id
            LEFT JOIN inventory_dimensions dim ON dim.inventory_entity_id = ie.id
            WHERE me.organization_id = $1
              AND ($2::int IS NULL OR me.project_map_id = $2)
            ORDER BY me.element_id
            """,
            current_user["organization_id"],
            project_map["id"] if project_map else None,
        )
        units = await connection.fetch(
            """
            SELECT
                ie.id,
                ie.entity_code AS code,
                ie.name,
                ie.entity_type AS type,
                ie.inventory_status AS status,
                ie.parent_id,
                ie.path,
                dim.saleable_area AS area,
                det.bhk_type AS bhk,
                pr.final_price AS price
            FROM inventory_entities ie
            LEFT JOIN inventory_dimensions dim ON dim.inventory_entity_id = ie.id
            LEFT JOIN inventory_details det ON det.inventory_entity_id = ie.id
            LEFT JOIN inventory_pricing pr ON pr.inventory_entity_id = ie.id
            WHERE ie.organization_id = $1
              AND ie.project_id = $2
              AND ie.entity_type IN ('flat', 'plot', 'villa', 'shop', 'office')
            ORDER BY ie.sort_order, ie.entity_code
            """,
            current_user["organization_id"],
            project_id,
        )
    return {
        "project": json_ready(project),
        "map": json_ready(project_map),
        "elements": json_ready(elements),
        "units": json_ready(units),
    }


@router.post("/maps", status_code=status.HTTP_201_CREATED)
async def create_map(
    payload: dict[str, Any] = Body(...),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "maps.create")
        await assert_project(connection, current_user["organization_id"], int(payload["project_id"]))
        data = clean_payload(payload, MAP_FIELDS)
        data["map_data"] = json.dumps(payload.get("map_data", {}))
        row = await connection.fetchrow(
            f"""
            INSERT INTO project_maps (organization_id, {", ".join(data.keys())}, created_by)
            VALUES ($1, {", ".join(f"${i}" for i in range(2, len(data) + 2))}, ${len(data) + 2})
            RETURNING *
            """,
            current_user["organization_id"],
            *data.values(),
            current_user["id"],
        )
    return json_ready(row)


@router.post("/projects-with-map", status_code=status.HTTP_201_CREATED)
async def create_project_with_map(
    payload: dict[str, Any] = Body(...),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    svg = str(payload.get("svg") or payload.get("map_data", {}).get("svg") or "").strip()
    if not svg:
        raise HTTPException(status_code=422, detail="SVG map is required")
    plot_ids = extract_svg_plot_ids(svg)
    if not plot_ids:
        raise HTTPException(status_code=422, detail="SVG must contain rect ids starting with plot_")

    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "projects.create")
        await ensure_permission(connection, current_user, "inventory.create")
        await ensure_permission(connection, current_user, "maps.create")
        data = clean_payload(payload, PROJECT_WITH_MAP_FIELDS)
        data.setdefault("project_type", "plotting")
        data.setdefault("status", "active")
        if not data.get("name"):
            raise HTTPException(status_code=422, detail="Project name is required")
        if not data.get("project_code"):
            base_code = re.sub(r"[^A-Z0-9]+", "-", str(data["name"]).upper()).strip("-")[:40] or "PROJECT"
            data["project_code"] = base_code
        supplied_project_code = bool(payload.get("project_code"))
        code_exists = await connection.fetchval(
            """
            SELECT EXISTS (
                SELECT 1 FROM projects WHERE organization_id = $1 AND project_code = $2
            )
            """,
            current_user["organization_id"],
            data["project_code"],
        )
        if code_exists and supplied_project_code:
            raise HTTPException(status_code=409, detail="Project code already exists")
        if code_exists:
            project_count = await connection.fetchval(
                "SELECT COUNT(*) + 1 FROM projects WHERE organization_id = $1",
                current_user["organization_id"],
            )
            data["project_code"] = f"{data['project_code']}-{int(project_count)}"[:50]

        async with connection.transaction():
            columns = ["organization_id", *data.keys()]
            project = await connection.fetchrow(
                f"""
                INSERT INTO projects ({", ".join(columns)})
                VALUES ({", ".join(f"${index}" for index in range(1, len(columns) + 1))})
                RETURNING *
                """,
                current_user["organization_id"],
                *data.values(),
            )
            project_code = project["project_code"] or f"PRJ-{project['id']}"
            project_map = await connection.fetchrow(
                """
                INSERT INTO project_maps (
                    organization_id, project_id, map_name, map_engine, version_no,
                    is_published, map_data, created_by
                )
                VALUES ($1, $2, $3, 'svg', 1, TRUE, $4::jsonb, $5)
                RETURNING *
                """,
                current_user["organization_id"],
                project["id"],
                payload.get("map_name") or f"{project['name']} Map",
                json.dumps({"svg": svg, "viewBox": extract_view_box(svg)}),
                current_user["id"],
            )

            created_plots = []
            created_floors = []
            for sort_order, element_id in enumerate(plot_ids, start=1):
                plot_code = re.sub(r"^plot_", "", element_id, flags=re.IGNORECASE).upper()
                plot = await connection.fetchrow(
                    """
                    INSERT INTO inventory_entities (
                        organization_id, project_id, entity_type, entity_code, name,
                        inventory_status, lifecycle_stage, level_no, path, sort_order, metadata
                    )
                    VALUES ($1, $2, 'plot', $3, $4, 'available', 'active_sales', 1, $5, $6, $7::jsonb)
                    RETURNING *
                    """,
                    current_user["organization_id"],
                    project["id"],
                    plot_code,
                    plot_code.replace("_", " "),
                    f"{project_code}/{plot_code}",
                    sort_order,
                    json.dumps({"svg_element_id": element_id}),
                )
                await connection.execute(
                    """
                    INSERT INTO inventory_dimensions (
                        organization_id, inventory_entity_id, saleable_area, measurement_unit
                    )
                    VALUES ($1, $2, 1000, 'sqyd')
                    """,
                    current_user["organization_id"],
                    plot["id"],
                )
                await connection.execute(
                    """
                    INSERT INTO inventory_pricing (
                        organization_id, inventory_entity_id, final_price, price_per_sqft, currency
                    )
                    VALUES ($1, $2, 5000000, 5000, 'INR')
                    """,
                    current_user["organization_id"],
                    plot["id"],
                )
                await connection.execute(
                    """
                    INSERT INTO map_elements (
                        organization_id, project_map_id, inventory_entity_id, element_id,
                        element_type, is_interactive, metadata
                    )
                    VALUES ($1, $2, $3, $4, 'plot', TRUE, $5::jsonb)
                    """,
                    current_user["organization_id"],
                    project_map["id"],
                    plot["id"],
                    element_id,
                    json.dumps({"generated_from_svg": True}),
                )
                created_plots.append(plot)

                for floor_no in range(1, 3):
                    floor_code = f"{plot_code}-F{floor_no}"
                    floor = await connection.fetchrow(
                        """
                        INSERT INTO inventory_entities (
                            organization_id, project_id, parent_id, entity_type, entity_code, name,
                            inventory_status, lifecycle_stage, level_no, path, sort_order
                        )
                        VALUES ($1, $2, $3, 'floor', $4, $5, 'available', 'active_sales', 2, $6, $7)
                        RETURNING *
                        """,
                        current_user["organization_id"],
                        project["id"],
                        plot["id"],
                        floor_code,
                        f"{plot_code} Floor {floor_no}",
                        f"{project_code}/{plot_code}/{floor_code}",
                        floor_no,
                    )
                    await connection.execute(
                        """
                        INSERT INTO inventory_dimensions (
                            organization_id, inventory_entity_id, saleable_area, measurement_unit
                        )
                        VALUES ($1, $2, 500, 'sqft')
                        """,
                        current_user["organization_id"],
                        floor["id"],
                    )
                    await connection.execute(
                        """
                        INSERT INTO inventory_pricing (
                            organization_id, inventory_entity_id, final_price, price_per_sqft, currency
                        )
                        VALUES ($1, $2, 2500000, 5000, 'INR')
                        """,
                        current_user["organization_id"],
                        floor["id"],
                    )
                    created_floors.append(floor)

            await audit_log(
                connection,
                organization_id=current_user["organization_id"],
                user_id=current_user["id"],
                entity_type="project",
                entity_id=project["id"],
                action="created_with_svg_map",
                new_data={
                    "project": json_ready(project),
                    "map": json_ready(project_map),
                    "plots": len(created_plots),
                    "floors": len(created_floors),
                },
            )

    return {
        "project": json_ready(project),
        "map": json_ready(project_map),
        "plots": json_ready(created_plots),
        "floors": json_ready(created_floors),
    }


@router.post("/map-elements", status_code=status.HTTP_201_CREATED)
async def create_map_element(
    payload: dict[str, Any] = Body(...),
    current_user: asyncpg.Record = Depends(get_current_user),
    pool: asyncpg.Pool = Depends(get_db_pool),
) -> dict[str, Any]:
    async with pool.acquire() as connection:
        await ensure_permission(connection, current_user, "maps.create")
        data = clean_payload(payload, ELEMENT_FIELDS)
        row = await connection.fetchrow(
            f"""
            INSERT INTO map_elements (organization_id, {", ".join(data.keys())})
            VALUES ($1, {", ".join(f"${i}" for i in range(2, len(data) + 2))})
            RETURNING *
            """,
            current_user["organization_id"],
            *data.values(),
        )
    return json_ready(row)
